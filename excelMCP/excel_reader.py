# Excel reader module

import pandas as pd
from openpyxl import load_workbook

class ExcelReader:
    """Class for reading Excel files with various options."""
    
    def __init__(self, file_path=None):
        """Initialize with optional file path."""
        self.file_path = file_path
        self.workbook = None
        self.data = None
    
    def load_file(self, file_path=None):
        """Load an Excel file."""
        if file_path:
            self.file_path = file_path
        
        if not self.file_path:
            raise ValueError("No file path provided")
            
        try:
            self.workbook = load_workbook(filename=self.file_path, data_only=True)
            return True
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def read_sheet_to_dataframe(self, sheet_name=None, sheet_index=0):
        """Read a specific sheet into a pandas DataFrame."""
        if not self.file_path:
            raise ValueError("No file path loaded. Call load_file() first.")
        
        try:
            if sheet_name:
                self.data = pd.read_excel(self.file_path, sheet_name=sheet_name)
            else:
                self.data = pd.read_excel(self.file_path, sheet_index=sheet_index)
            return self.data
        except Exception as e:
            print(f"Error reading sheet to dataframe: {e}")
            return None
    
    def get_sheet_names(self):
        """Get all sheet names from the workbook."""
        if not self.workbook:
            if not self.load_file():
                return []
        
        return self.workbook.sheetnames
