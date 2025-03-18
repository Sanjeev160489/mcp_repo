# Excel writer module

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelWriter:
    """Class for writing data to Excel files with formatting options."""
    
    def __init__(self, file_path=None):
        """Initialize with optional file path."""
        self.file_path = file_path
        self.workbook = Workbook()
    
    def set_file_path(self, file_path):
        """Set the file path for the output Excel file."""
        self.file_path = file_path
    
    def dataframe_to_excel(self, df, sheet_name='Sheet1', start_cell='A1', include_header=True):
        """Write a pandas DataFrame to an Excel sheet."""
        if not isinstance(df, pd.DataFrame):
            raise TypeError("Data must be a pandas DataFrame")
        
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(sheet_name)
        else:
            sheet = self.workbook[sheet_name]
        
        # Extract row and column from start_cell
        col_letter = ''.join(filter(str.isalpha, start_cell))
        row_num = int(''.join(filter(str.isdigit, start_cell)))
        
        # Write header if requested
        if include_header:
            for col_idx, col_name in enumerate(df.columns):
                cell = sheet.cell(row=row_num, column=col_idx + 1)
                cell.value = col_name
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            row_num += 1
        
        # Write data
        for row_idx, row in enumerate(df.values):
            for col_idx, value in enumerate(row):
                sheet.cell(row=row_idx + row_num, column=col_idx + 1).value = value
    
    def save(self, file_path=None):
        """Save the workbook to the specified file path."""
        if file_path:
            self.file_path = file_path
        
        if not self.file_path:
            raise ValueError("No file path provided for saving")
        
        try:
            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False
